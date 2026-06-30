#!/usr/bin/env python3
"""
Genera un script SQL para migrar el histórico del Excel REPOSITORIO (hoja
'REPOSITORIO') al esquema del sistema de dosímetros.

Uso:
    python3 generar_migracion.py <archivo.xlsx> [salida.sql]

Requisitos: pip install openpyxl

Supuestos de la migración (revisar y ajustar si es necesario):
- Empresa: todas las asignaciones se marcan como 'Photomat'.
- Catálogos (empresa, tipo_dosimetro, tipo_porta base) ya existen por las
  migraciones Flyway V1/V2; se referencian por nombre con variables SQL.
- Portas que no están en el catálogo se crean: 'Sin armar' (TLD),
  'Cristales' (Cristal) y 'OSL con broche' (OSL).
- El tipo de dosímetro se deduce del tipo de porta.
- Estado del dosímetro: 'asignado' si su última fila tiene cliente; si no,
  'disponible'. La ubicación actual proviene de su última fila.
- Filas sin cliente NO generan asignación (el dosímetro igual se crea).
- Filas con cliente pero sin ejecutivo usan el ejecutivo 'Sin asignar'.
- 'FECHA DE INGRESO' del Excel se guarda como link_trello.
- Las tablas migradas (asignacion, dosimetro, tarea, cliente, ejecutivo) se
  vacían al inicio (TRUNCATE); los catálogos NO se tocan.
"""

import sys
import openpyxl
from datetime import datetime, date

# Columnas de la hoja REPOSITORIO
C_FECHA_ING, C_NUM, C_EJEC, C_CLI, C_TRI, C_FASIG, C_PORTA, C_BAND, C_SLOT, C_TAREA = range(10)

# Mapeo de porta del archivo -> (nombre en catálogo, tipo de dosímetro)
PORTA_MAP = {
    'PORTA GRINGO':   ('Porta gringo', 'TLD'),
    'PORTA VIEJO':    ('Porta viejo', 'TLD'),
    'OSL':            ('OSL', 'OSL'),
    'OSL CON BROCHE': ('OSL con broche', 'OSL'),
    'SIN ARMAR':      ('Sin armar', 'TLD'),
    'ANILLO':         ('Anillo', 'Cristal'),
    'PULSERA':        ('Pulsera', 'Cristal'),
    'CRISTALINO':     ('Cristalino', 'Cristal'),
    'CRISTALES':      ('Cristales', 'Cristal'),
}
PORTA_DEFAULT = ('Sin armar', 'TLD')

# Portas a crear si no existen (nombre -> tipo)
PORTAS_NUEVAS = {
    'Sin armar': 'TLD',
    'Cristales': 'Cristal',
    'OSL con broche': 'OSL',
}

TD_VAR = {'OSL': '@td_osl', 'TLD': '@td_tld', 'Cristal': '@td_cristal'}


def sstr(v):
    return '' if v is None else str(v).strip()


def esc(v):
    if v is None:
        return 'NULL'
    s = str(v).replace('\\', '\\\\').replace("'", "''").strip()
    return "'" + s + "'"


def porta_var(nombre):
    # nombre de catálogo -> variable SQL
    return '@p_' + (nombre.lower()
                    .replace(' ', '_').replace('(', '').replace(')', '')
                    .replace('á', 'a'))


def to_date(v):
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    return None


def main():
    if len(sys.argv) < 2:
        print('Uso: python3 generar_migracion.py <archivo.xlsx> [salida.sql]')
        sys.exit(1)
    ruta = sys.argv[1]
    salida = sys.argv[2] if len(sys.argv) > 2 else 'migracion_repositorio.sql'

    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb['REPOSITORIO'] if 'REPOSITORIO' in wb.sheetnames else wb.worksheets[0]

    it = ws.iter_rows(values_only=True)
    next(it)  # encabezado

    ejecutivos = {}      # nombre normalizado -> id
    clientes = {}        # nombre -> id
    tareas = {}          # numero_tarea -> id
    tarea_fecha = {}     # numero_tarea -> fecha mínima
    cliente_ejec = {}    # cliente -> ejecutivo (para responsable)
    dosimetros = {}      # numdosim -> dict(rows=[...])
    filas = []           # filas crudas normalizadas
    descartadas = [0]    # filas con número no numérico

    EJEC_PLACEHOLDER = 'Sin asignar'

    for r in it:
        if r is None:
            continue
        row = list(r) + [None] * (10 - len(r))
        num = row[C_NUM]
        cli = sstr(row[C_CLI])
        ej = sstr(row[C_EJEC]).title()
        if num is None and not cli and not ej:
            continue
        if num is None:
            continue  # sin número no podemos crear dosímetro
        try:
            num = int(float(num))
        except (ValueError, TypeError):
            descartadas[0] += 1
            continue  # número no numérico (ej. 'EEEE')
        porta_raw = sstr(row[C_PORTA]).upper()
        cat_nombre, tipo = PORTA_MAP.get(porta_raw, PORTA_DEFAULT)
        # Trimestre debe ser texto tipo '2T2025'; descartar valores no válidos.
        tri = sstr(row[C_TRI])
        if 'T' not in tri.upper():
            tri = ''
        fasig = to_date(row[C_FASIG])
        tarea_num = None if row[C_TAREA] is None else sstr(row[C_TAREA])
        band = row[C_BAND] if isinstance(row[C_BAND], (int, float)) else None
        slot = row[C_SLOT] if isinstance(row[C_SLOT], (int, float)) else None
        link = sstr(row[C_FECHA_ING]) or None

        fila = {
            'num': num, 'cli': cli, 'ej': ej, 'tri': tri, 'fasig': fasig,
            'porta': cat_nombre, 'tipo': tipo, 'tarea': tarea_num,
            'band': int(band) if band is not None else None,
            'slot': int(slot) if slot is not None else None,
            'link': link,
        }
        filas.append(fila)

        if ej:
            ejecutivos.setdefault(ej, None)
        if cli:
            clientes.setdefault(cli, None)
            cliente_ejec[cli] = ej or cliente_ejec.get(cli)
        if tarea_num:
            tareas.setdefault(tarea_num, None)
            if fasig and (tarea_num not in tarea_fecha or fasig < tarea_fecha[tarea_num]):
                tarea_fecha[tarea_num] = fasig
        dosimetros.setdefault(num, []).append(fila)

    # Asignar ids
    ejecutivos.setdefault(EJEC_PLACEHOLDER, None)
    for i, k in enumerate(ejecutivos, start=1):
        ejecutivos[k] = i
    for i, k in enumerate(clientes, start=1):
        clientes[k] = i
    for i, k in enumerate(tareas, start=1):
        tareas[k] = i
    dosim_id = {num: i for i, num in enumerate(dosimetros, start=1)}

    portas_usadas = sorted({f['porta'] for f in filas} | set(PORTAS_NUEVAS))

    out = []
    w = out.append
    w('-- Migración histórica del REPOSITORIO (generado automáticamente)')
    w('SET NAMES utf8mb4;')
    w('SET FOREIGN_KEY_CHECKS = 0;')
    w('TRUNCATE TABLE ASIGNACION;')
    w('TRUNCATE TABLE dosimetro;')
    w('TRUNCATE TABLE tarea;')
    w('TRUNCATE TABLE cliente;')
    w('TRUNCATE TABLE ejecutivo;')
    w('SET FOREIGN_KEY_CHECKS = 1;')
    w('')
    w("SET @empresa := (SELECT id FROM empresa WHERE nombre = 'Photomat' LIMIT 1);")
    w("SET @td_osl := (SELECT id FROM tipo_dosimetro WHERE nombre = 'OSL' LIMIT 1);")
    w("SET @td_tld := (SELECT id FROM tipo_dosimetro WHERE nombre = 'TLD' LIMIT 1);")
    w("SET @td_cristal := (SELECT id FROM tipo_dosimetro WHERE nombre = 'Cristal' LIMIT 1);")
    w('')
    # Crear portas faltantes de forma idempotente
    for nombre, tipo in PORTAS_NUEVAS.items():
        w(f"INSERT INTO tipo_porta (nombre, tipo_dosimetro_id) "
          f"SELECT {esc(nombre)}, {TD_VAR[tipo]} FROM dual "
          f"WHERE NOT EXISTS (SELECT 1 FROM tipo_porta WHERE nombre = {esc(nombre)});")
    w('')
    # Variables de portas
    for nombre in portas_usadas:
        w(f"SET {porta_var(nombre)} := (SELECT id FROM tipo_porta WHERE nombre = {esc(nombre)} LIMIT 1);")
    w('')

    def batched_insert(header, valores, batch=500):
        for i in range(0, len(valores), batch):
            chunk = valores[i:i + batch]
            w(header)
            w(',\n'.join(chunk) + ';')

    # Ejecutivos
    vals = [f"({eid}, {esc(nombre)}, NULL, 1)" for nombre, eid in ejecutivos.items()]
    batched_insert('INSERT INTO ejecutivo (id, nombre, email, activo) VALUES', vals)
    w('')

    # Clientes (con ejecutivo responsable)
    vals = []
    for nombre, cid in clientes.items():
        ej = cliente_ejec.get(nombre)
        ej_id = ejecutivos.get(ej) if ej else None
        vals.append(f"({cid}, {esc(nombre)}, NULL, {ej_id if ej_id else 'NULL'}, 1)")
    batched_insert('INSERT INTO cliente (id, razon_social, nombre_corto, ejecutivo_id, activo) VALUES', vals)
    w('')

    # Tareas
    vals = []
    for numt, tid in tareas.items():
        f = tarea_fecha.get(numt)
        vals.append(f"({tid}, {esc(numt)}, {esc(f) if f else 'CURDATE()'}, NULL)")
    batched_insert('INSERT INTO tarea (id, numero_tarea, fecha_creacion, observacion) VALUES', vals)
    w('')

    # Dosímetros (ubicación/estado actual = última fila)
    def fila_key(f):
        return (f['fasig'] or '0000-00-00')

    vals = []
    for num, rows in dosimetros.items():
        ult = sorted(rows, key=fila_key)[-1]
        did = dosim_id[num]
        td = TD_VAR[ult['tipo']]
        pv = porta_var(ult['porta'])
        tarea_id = tareas.get(ult['tarea']) if ult['tarea'] else None
        estado = 'asignado' if ult['cli'] else 'disponible'
        vals.append(
            f"({did}, {num}, {td}, {pv}, "
            f"{tarea_id if tarea_id else 'NULL'}, "
            f"{ult['band'] if ult['band'] is not None else 'NULL'}, "
            f"{ult['slot'] if ult['slot'] is not None else 'NULL'}, "
            f"'{estado}', NULL)"
        )
    batched_insert(
        'INSERT INTO dosimetro (id, numero, tipo_dosimetro_id, tipo_porta_id, '
        'tarea_id, numero_bandeja, slot_bandeja, estado, observacion) VALUES', vals)
    w('')

    # Asignaciones (solo filas con cliente)
    vals = []
    n_asig = 0
    for f in filas:
        if not f['cli']:
            continue
        if not f['fasig'] or not f['tri']:
            continue  # requeridos NOT NULL
        did = dosim_id[f['num']]
        cid = clientes[f['cli']]
        eid = ejecutivos.get(f['ej']) or ejecutivos[EJEC_PLACEHOLDER]
        pv = porta_var(f['porta'])
        tarea_id = tareas.get(f['tarea']) if f['tarea'] else None
        vals.append(
            f"({did}, {cid}, {eid}, @empresa, {pv}, "
            f"{tarea_id if tarea_id else 'NULL'}, "
            f"{f['band'] if f['band'] is not None else 'NULL'}, "
            f"{f['slot'] if f['slot'] is not None else 'NULL'}, "
            f"{esc(f['tri'])}, {esc(f['fasig'])}, {esc(f['link'])})"
        )
        n_asig += 1
    batched_insert(
        'INSERT INTO ASIGNACION (dosimetro_id, cliente_id, ejecutivo_id, empresa_id, '
        'tipo_porta_id, tarea_id, numero_bandeja, slot_bandeja, trimestre, '
        'fecha_asignacion, link_trello) VALUES', vals)

    with open(salida, 'w', encoding='utf-8') as fh:
        fh.write('\n'.join(out) + '\n')

    print(f'Generado: {salida}')
    print(f'  Ejecutivos: {len(ejecutivos)}')
    print(f'  Clientes:   {len(clientes)}')
    print(f'  Tareas:     {len(tareas)}')
    print(f'  Dosímetros: {len(dosimetros)}')
    print(f'  Asignaciones: {n_asig}')
    print(f'  Filas descartadas (número no numérico): {descartadas[0]}')


if __name__ == '__main__':
    main()
